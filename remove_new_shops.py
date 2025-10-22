#!/usr/bin/env python3
"""
Remove Recently Added Shops from shops_details.json

This script removes all shops that were added from new-shop-list-updated.xlsx (highlighted shops)
while preserving shops from old-shop-list-updated.xlsx
"""

import json
import openpyxl
import sys
from datetime import datetime

def safe_str(value):
    """Convert value to string safely"""
    if value is None:
        return ""
    return str(value).strip()

def get_highlighted_licenses():
    """Get list of license numbers from highlighted rows in new-shop-list-updated.xlsx"""
    print("=" * 80)
    print("LOADING HIGHLIGHTED SHOPS FROM new-shop-list-updated.xlsx")
    print("=" * 80)
    
    wb = openpyxl.load_workbook('new-shop-list-updated.xlsx')
    ws = wb.active
    
    highlighted_licenses = set()
    
    for row_idx in range(2, ws.max_row + 1):
        # Check if row is highlighted (yellow background)
        cell = ws.cell(row_idx, 1)
        is_highlighted = False
        if cell.fill and cell.fill.patternType and cell.fill.patternType != 'none':
            fill_color = cell.fill.fgColor
            if fill_color and hasattr(fill_color, 'rgb') and fill_color.rgb == 'FFFFFF00':
                is_highlighted = True
        
        # Only process highlighted rows
        if not is_highlighted:
            continue
        
        license_no = safe_str(ws.cell(row_idx, 1).value)
        
        if license_no:
            highlighted_licenses.add(license_no)
    
    print(f"   ✓ Found {len(highlighted_licenses)} highlighted licenses to remove")
    return highlighted_licenses

def main():
    print("\n" + "=" * 80)
    print("REMOVING RECENTLY ADDED SHOPS FROM shops_details.json")
    print("=" * 80)
    print()
    
    # Load shops_details.json
    print("[1/5] Loading shops_details.json...")
    with open('shops_details.json', 'r', encoding='utf-8') as f:
        shops_details = json.load(f)
    print(f"   ✓ Loaded {len(shops_details)} shops")
    print()
    
    # Get highlighted licenses from Excel
    print("[2/5] Identifying shops to remove...")
    highlighted_licenses = get_highlighted_licenses()
    print()
    
    # Create backup
    backup_filename = f'shops_details.json.backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
    print(f"[3/5] Creating backup: {backup_filename}")
    with open(backup_filename, 'w', encoding='utf-8') as f:
        json.dump(shops_details, f, ensure_ascii=False, indent=2)
    print("   ✓ Backup created")
    print()
    
    # Remove shops with licenses from highlighted list
    print("[4/5] Removing shops from shops_details.json...")
    shops_to_remove = []
    
    for shop_name, shop_data in shops_details.items():
        license_no = shop_data.get('licenseNo')
        if license_no and license_no in highlighted_licenses:
            shops_to_remove.append(shop_name)
    
    print(f"   ✓ Found {len(shops_to_remove)} shops to remove")
    
    # Remove the shops
    for shop_name in shops_to_remove:
        del shops_details[shop_name]
        print(f"   - Removed: {shop_name}")
    
    print()
    
    # Save updated shops_details.json
    print("[5/5] Saving updated shops_details.json...")
    with open('shops_details.json', 'w', encoding='utf-8') as f:
        json.dump(shops_details, f, ensure_ascii=False, indent=2)
    print("   ✓ Saved successfully")
    
    print()
    print("=" * 80)
    print("✅ REMOVAL COMPLETED SUCCESSFULLY!")
    print("=" * 80)
    print(f"\nStatistics:")
    print(f"  - Original shop count: {len(shops_details) + len(shops_to_remove)}")
    print(f"  - Shops removed: {len(shops_to_remove)}")
    print(f"  - Remaining shop count: {len(shops_details)}")
    print(f"\nBackup saved as: {backup_filename}")
    
    return True

if __name__ == '__main__':
    try:
        success = main()
        sys.exit(0 if success else 1)
    except Exception as e:
        print(f"\n❌ ERROR: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)
