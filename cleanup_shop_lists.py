#!/usr/bin/env python3
"""
Script to clean up shop list files by removing non-highlighted shops.

This script:
1. In NEW file: Keeps only yellow-highlighted shops, removes all others
2. In OLD file: Keeps all shops (no filtering based on highlighting)
"""

import openpyxl
from openpyxl.styles import PatternFill
import shutil
from datetime import datetime

def is_yellow_highlighted(cell):
    """Check if a cell has yellow highlighting"""
    if cell.fill and cell.fill.fgColor:
        if hasattr(cell.fill.fgColor, 'rgb') and cell.fill.fgColor.rgb:
            color = str(cell.fill.fgColor.rgb).upper()
            # Check for yellow color (FFFFFF00 and variations)
            if 'FFFF' in color and color != '00000000':
                return True
    return False

def backup_file(filename):
    """Create a backup of the file"""
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    backup_name = filename.replace('.xlsx', f'_backup_{timestamp}.xlsx')
    shutil.copy2(filename, backup_name)
    print(f'Backup created: {backup_name}')
    return backup_name

def clean_new_shop_list(filename='new-shop-list-updated.xlsx'):
    """
    Clean new shop list by keeping only yellow-highlighted rows
    """
    print(f'\n=== Processing {filename} ===')
    
    # Create backup first
    backup_file(filename)
    
    # Load workbook
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    total_rows = ws.max_row - 1  # Excluding header
    print(f'Total shops before cleanup: {total_rows}')
    
    # Identify rows to keep (yellow highlighted)
    rows_to_keep = [1]  # Always keep header
    yellow_count = 0
    
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row_idx, 1)
        if is_yellow_highlighted(cell):
            rows_to_keep.append(row_idx)
            yellow_count += 1
    
    print(f'Yellow highlighted shops to keep: {yellow_count}')
    print(f'Non-highlighted shops to delete: {total_rows - yellow_count}')
    
    # Create new workbook with only the rows to keep
    new_wb = openpyxl.Workbook()
    new_ws = new_wb.active
    new_ws.title = ws.title
    
    # Copy rows to keep
    new_row = 1
    for old_row in rows_to_keep:
        for col in range(1, ws.max_column + 1):
            old_cell = ws.cell(old_row, col)
            new_cell = new_ws.cell(new_row, col)
            
            # Copy value
            new_cell.value = old_cell.value
            
            # Copy formatting
            if old_cell.has_style:
                new_cell.font = old_cell.font.copy()
                new_cell.border = old_cell.border.copy()
                new_cell.fill = old_cell.fill.copy()
                new_cell.number_format = old_cell.number_format
                new_cell.protection = old_cell.protection.copy()
                new_cell.alignment = old_cell.alignment.copy()
        
        new_row += 1
    
    # Copy column widths
    for col in range(1, ws.max_column + 1):
        column_letter = openpyxl.utils.get_column_letter(col)
        if column_letter in ws.column_dimensions:
            new_ws.column_dimensions[column_letter].width = ws.column_dimensions[column_letter].width
    
    # Save the cleaned file
    new_wb.save(filename)
    print(f'Cleaned file saved: {filename}')
    print(f'Total shops after cleanup: {yellow_count}')
    
    return yellow_count

def clean_old_shop_list(filename='old-shop-list-updated.xlsx'):
    """
    Clean old shop list - for now, just keep all shops
    Based on analysis, there are no yellow highlights in this file
    """
    print(f'\n=== Processing {filename} ===')
    
    # Create backup first
    backup_file(filename)
    
    # Load workbook
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    total_rows = ws.max_row - 1  # Excluding header
    print(f'Total shops in file: {total_rows}')
    print(f'No yellow highlighting found, keeping all shops')
    
    # No changes needed for old file based on analysis
    # Just save it as is (backup already created)
    print(f'File unchanged: {filename}')
    
    return total_rows

def main():
    """Main cleanup function"""
    print('=' * 60)
    print('Shop List Cleanup Script')
    print('=' * 60)
    
    # Clean new shop list (remove non-yellow shops)
    new_kept = clean_new_shop_list()
    
    # Old shop list (keep all)
    old_kept = clean_old_shop_list()
    
    print('\n' + '=' * 60)
    print('CLEANUP SUMMARY')
    print('=' * 60)
    print(f'New shop list: {new_kept} shops remaining (only yellow-highlighted)')
    print(f'Old shop list: {old_kept} shops remaining (all kept)')
    print(f'Total shops across both files: {new_kept + old_kept}')
    print('=' * 60)

if __name__ == '__main__':
    main()
