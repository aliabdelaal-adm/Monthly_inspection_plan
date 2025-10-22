#!/usr/bin/env python3
"""
Merge New Shops from Excel into shops_details.json
This script integrates new shop data from new-shop-list-updated.xlsx that are not yet in the system.
"""

import json
import openpyxl
import sys
from datetime import datetime

def safe_str(value):
    """Convert value to string safely"""
    if value is None:
        return ""
    return str(value).strip()

def generate_adm_code(existing_codes):
    """Generate next available ADM code"""
    # Extract numeric parts from existing codes
    max_num = 0
    for code in existing_codes:
        if code and code.startswith('ADM'):
            try:
                num = int(code[3:])
                max_num = max(max_num, num)
            except ValueError:
                continue
    
    # Return next number
    return f"ADM{str(max_num + 1).zfill(4)}"

def main():
    print("=" * 80)
    print("MERGING NEW SHOPS FROM EXCEL INTO shops_details.json")
    print("=" * 80)
    
    # Load existing shops_details.json
    print("\n[1/6] Loading shops_details.json...")
    with open('shops_details.json', 'r', encoding='utf-8') as f:
        shops_details = json.load(f)
    print(f"   ✓ Loaded {len(shops_details)} existing shops")
    
    # Track existing licenses and ADM codes
    existing_licenses = set()
    existing_adm_codes = set()
    
    for shop_name, shop_data in shops_details.items():
        if shop_data.get('licenseNo'):
            existing_licenses.add(shop_data['licenseNo'])
        if shop_data.get('admCode'):
            existing_adm_codes.add(shop_data['admCode'])
    
    # Create backup
    backup_filename = f'shops_details.json.backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}'
    print(f"\n[2/6] Creating backup: {backup_filename}")
    with open(backup_filename, 'w', encoding='utf-8') as f:
        json.dump(shops_details, f, ensure_ascii=False, indent=2)
    print("   ✓ Backup created")
    
    # Load Excel data
    print("\n[3/6] Loading Excel data from new-shop-list-updated.xlsx...")
    wb = openpyxl.load_workbook('new-shop-list-updated.xlsx')
    ws = wb.active
    print(f"   ✓ Loaded Excel file with {ws.max_row} rows")
    
    # Find new shops
    print("\n[4/6] Identifying new shops...")
    new_shops = []
    
    for row_idx in range(2, ws.max_row + 1):
        license_no = safe_str(ws.cell(row_idx, 1).value)
        name_ar = safe_str(ws.cell(row_idx, 2).value)
        name_en = safe_str(ws.cell(row_idx, 3).value)
        
        if license_no and (name_ar or name_en):
            # Check if this shop is new
            if license_no not in existing_licenses:
                shop_data = {
                    'licenseNo': license_no,
                    'nameAr': name_ar,
                    'nameEn': name_en,
                    'address': safe_str(ws.cell(row_idx, 17).value),
                    'contact': safe_str(ws.cell(row_idx, 12).value),
                    'email': safe_str(ws.cell(row_idx, 13).value),
                    'activity': safe_str(ws.cell(row_idx, 19).value),
                }
                new_shops.append(shop_data)
    
    print(f"   ✓ Found {len(new_shops)} new shops to add")
    
    if len(new_shops) == 0:
        print("\n✓ No new shops to add. All shops are already in the system.")
        return True
    
    # Add new shops to shops_details
    print("\n[5/6] Adding new shops to shops_details.json...")
    for shop in new_shops:
        # Use Arabic name as key if available, otherwise English name
        shop_key = shop['nameAr'] if shop['nameAr'] else shop['nameEn']
        
        # If shop name already exists, append license number to make it unique
        if shop_key in shops_details:
            shop_key = f"{shop_key} - {shop['licenseNo']}"
            print(f"   ⚠ Name conflict detected, using unique key: {shop_key}")
        
        # Generate ADM code
        adm_code = generate_adm_code(existing_adm_codes)
        existing_adm_codes.add(adm_code)
        
        # Build shop entry
        shops_details[shop_key] = {
            'nameAr': shop['nameAr'],
            'nameEn': shop['nameEn'],
            'licenseNo': shop['licenseNo'],
            'locationMap': '',
            'admCode': adm_code,
            'address': shop['address'],
            'contact': shop['contact'],
            'email': shop['email'],
            'activity': shop['activity']
        }
        
        print(f"   ✓ Added: {shop_key} ({adm_code})")
    
    # Save updated shops_details.json
    print("\n[6/6] Saving updated shops_details.json...")
    with open('shops_details.json', 'w', encoding='utf-8') as f:
        json.dump(shops_details, f, ensure_ascii=False, indent=2)
    print("   ✓ Saved successfully")
    
    print("\n" + "=" * 80)
    print("✅ MERGE COMPLETED SUCCESSFULLY!")
    print("=" * 80)
    print(f"\nStatistics:")
    print(f"  - Original shop count: {len(shops_details) - len(new_shops)}")
    print(f"  - New shops added: {len(new_shops)}")
    print(f"  - Total shop count: {len(shops_details)}")
    print(f"\nBackup saved as: {backup_filename}")
    
    return True

if __name__ == '__main__':
    try:
        success = main()
        sys.exit(0 if success else 1)
    except Exception as e:
        print(f"\n❌ ERROR: {e}", file=sys.stderr)
        import traceback
        traceback.print_exc()
        sys.exit(1)
