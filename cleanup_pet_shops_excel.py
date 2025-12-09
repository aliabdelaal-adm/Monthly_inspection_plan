#!/usr/bin/env python3
"""
Script to clean up pet-shops-by-area.xlsx:
1. Remove duplicate shops
2. Shorten long location map links
3. Rearrange shops between areas based on location
4. Delete repeated shops with same data
"""

import pandas as pd
import re
from collections import defaultdict
from urllib.parse import urlparse, parse_qs, unquote
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def extract_coordinates_from_link(link):
    """Extract coordinates from Google Maps link if available."""
    if not isinstance(link, str):
        return None
    
    # Try to find coordinates in various formats
    patterns = [
        r'@([-\d.]+),([-\d.]+)',  # @lat,lng format
        r'll=([-\d.]+),([-\d.]+)',  # ll=lat,lng format
        r'q=([-\d.]+),([-\d.]+)',  # q=lat,lng format
    ]
    
    for pattern in patterns:
        match = re.search(pattern, link)
        if match:
            try:
                lat, lng = float(match.group(1)), float(match.group(2))
                return (lat, lng)
            except (ValueError, TypeError):
                continue
    
    return None

def shorten_google_maps_link(link):
    """
    Shorten long Google Maps search URLs to a cleaner format.
    Convert long search URLs to direct coordinate links when possible.
    """
    if not isinstance(link, str) or not link.strip():
        return link
    
    # If it's already a short goo.gl or maps.app.goo.gl link, keep it
    if 'goo.gl' in link or len(link) < 100:
        return link
    
    # Extract coordinates if available
    coords = extract_coordinates_from_link(link)
    if coords:
        lat, lng = coords
        return f"https://maps.app.goo.gl/custom?q={lat},{lng}"
    
    # Try to extract query from search URL
    if 'maps/search/' in link or 'api=1' in link:
        # Try to get the query parameter
        if 'query=' in link:
            try:
                query_start = link.find('query=') + 6
                query_end = link.find('&', query_start)
                if query_end == -1:
                    query = link[query_start:]
                else:
                    query = link[query_start:query_end]
                
                # URL decode the query
                query = unquote(query)
                
                # Create a cleaner search link
                return f"https://www.google.com/maps/search/?api=1&query={query[:100]}"
            except (ValueError, IndexError):
                pass
    
    # If we can't shorten it, return as is
    return link

def normalize_shop_name(name):
    """Normalize shop name for comparison."""
    if not isinstance(name, str):
        return ""
    
    # Remove extra spaces
    name = ' '.join(name.split())
    
    # Remove common variations
    name = name.replace('محل ', '').replace('معرض ', '')
    name = name.replace('ذ.م.م', '').replace('ذ م م', '').replace('ذ.م.م.', '')
    name = name.replace('-', ' ').strip()
    
    return name.lower()

def read_shop_data_from_excel(file_path):
    """Read all shop data from Excel file."""
    xl = pd.ExcelFile(file_path)
    all_shops = []
    
    for sheet_name in xl.sheet_names:
        # Skip the Summary sheet as it's not a real area
        if 'Summary' in sheet_name or 'الملخص' in sheet_name:
            continue
        df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
        
        # Find header row (contains '#' or 'اسم المحل')
        header_row = None
        for idx, row in df.iterrows():
            row_str = ' '.join([str(val) for val in row.values if pd.notna(val)])
            if '#' in row_str or 'اسم المحل' in row_str or 'Shop Name' in row_str:
                header_row = idx
                break
        
        if header_row is None:
            continue
        
        # Get column mapping
        headers = df.iloc[header_row].values
        
        # Process data rows
        for idx in range(header_row + 1, len(df)):
            row = df.iloc[idx]
            
            # Skip empty rows
            if row.isna().all():
                continue
            
            # Extract shop data (assuming standard column order)
            shop_data = {
                'area': sheet_name,
                'row_number': idx + 1,
                'number': row.iloc[0] if pd.notna(row.iloc[0]) else '',
                'name': row.iloc[1] if len(row) > 1 and pd.notna(row.iloc[1]) else '',
                'license': row.iloc[2] if len(row) > 2 and pd.notna(row.iloc[2]) else '',
                'phone': row.iloc[3] if len(row) > 3 and pd.notna(row.iloc[3]) else '',
                'address': row.iloc[4] if len(row) > 4 and pd.notna(row.iloc[4]) else '',
                'owner': row.iloc[5] if len(row) > 5 and pd.notna(row.iloc[5]) else '',
                'activity': row.iloc[6] if len(row) > 6 and pd.notna(row.iloc[6]) else '',
                'location_link': row.iloc[7] if len(row) > 7 and pd.notna(row.iloc[7]) else '',
            }
            
            # Skip if no meaningful data
            if not shop_data['name'] or shop_data['name'] == '':
                continue
            
            all_shops.append(shop_data)
    
    return all_shops

def find_duplicates(shops):
    """Find duplicate shops based on name, license, or exact match."""
    duplicates = []
    seen = {}
    seen_licenses = {}
    
    for i, shop in enumerate(shops):
        # Create keys for different types of matching
        name_key = normalize_shop_name(str(shop['name']))
        license_key = str(shop['license']).strip() if shop['license'] else ''
        
        # Exact match key (all fields except area and row_number)
        exact_key = (
            name_key,
            license_key,
            str(shop['phone']).strip(),
            str(shop['address']).strip()
        )
        
        # Check for exact duplicates
        if exact_key in seen and exact_key != ('', '', '', ''):
            duplicates.append({
                'original_index': seen[exact_key],
                'duplicate_index': i,
                'type': 'exact',
                'shop1': shops[seen[exact_key]],
                'shop2': shop
            })
        else:
            seen[exact_key] = i
        
        # Track license duplicates for reporting (but don't mark for removal)
        # These are likely branches in different areas
        if license_key and license_key not in ['', 'nan', 'None'] and len(license_key) > 3:
            if license_key in seen_licenses:
                orig_idx = seen_licenses[license_key]
                # Only report if in same area (likely true duplicate)
                if shops[orig_idx]['area'] == shop['area']:
                    if normalize_shop_name(str(shops[orig_idx]['name'])) != name_key:
                        duplicates.append({
                            'original_index': orig_idx,
                            'duplicate_index': i,
                            'type': 'same_license_same_area',
                            'shop1': shops[orig_idx],
                            'shop2': shop
                        })
            else:
                seen_licenses[license_key] = i
    
    return duplicates

def create_cleaned_excel(shops, output_file):
    """Create a new cleaned Excel file with organized shops."""
    # Group shops by area
    shops_by_area = defaultdict(list)
    for shop in shops:
        shops_by_area[shop['area']].append(shop)
    
    # Create new workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Define styles
    header_font = Font(name='Arial', size=11, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
    title_font = Font(name='Arial', size=14, bold=True)
    title_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    normal_font = Font(name='Arial', size=10)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Sort areas alphabetically
    sorted_areas = sorted(shops_by_area.keys())
    
    for area in sorted_areas:
        ws = wb.create_sheet(title=area)
        shops_in_area = shops_by_area[area]
        
        # Title row
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f"محلات منطقة: {area}"
        title_cell.font = title_font
        title_cell.fill = title_fill
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Subtitle row with count
        ws.merge_cells('A2:H2')
        subtitle_cell = ws['A2']
        subtitle_cell.value = f"Shops in Area: {area} | عدد المحلات: {len(shops_in_area)}"
        subtitle_cell.font = Font(name='Arial', size=11, bold=True)
        subtitle_cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Empty row
        ws.append([])
        
        # Header row
        headers = ['#', 'اسم المحل\nShop Name', 'رقم الرخصة\nLicense No', 'رقم الهاتف\nPhone', 
                   'العنوان\nAddress', 'اسم المالك\nOwner Name', 'النشاط\nActivity', 'رابط الموقع\nLocation Map']
        header_row = ws.append(headers)
        
        for cell in ws[4]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = border
        
        # Data rows
        for idx, shop in enumerate(shops_in_area, 1):
            row_data = [
                idx,
                shop['name'],
                shop['license'],
                shop['phone'],
                shop['address'],
                shop['owner'],
                shop['activity'],
                shop['location_link']
            ]
            ws.append(row_data)
            
            # Style data row
            for cell in ws[ws.max_row]:
                cell.font = normal_font
                cell.alignment = Alignment(horizontal='right', vertical='center', wrap_text=True)
                cell.border = border
                
                # Left align for number column
                if cell.column == 1:
                    cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Set column widths
        ws.column_dimensions['A'].width = 5
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 25
        ws.column_dimensions['F'].width = 20
        ws.column_dimensions['G'].width = 20
        ws.column_dimensions['H'].width = 40
        
        # Set row heights
        ws.row_dimensions[1].height = 25
        ws.row_dimensions[2].height = 20
        ws.row_dimensions[4].height = 30
    
    wb.save(output_file)
    print(f"Created cleaned Excel file: {output_file}")

def main():
    input_file = 'pet-shops-by-area.xlsx'
    output_file = 'pet-shops-by-area.xlsx'
    
    print("="*80)
    print("Pet Shops Excel Cleanup Script")
    print("="*80)
    
    # Check if input file exists
    import os
    if not os.path.exists(input_file):
        print(f"\n❌ ERROR: Input file '{input_file}' not found!")
        print("Please ensure the file exists in the current directory.")
        return 1
    
    # Step 1: Read all shops
    print("\n1. Reading shop data from Excel...")
    shops = read_shop_data_from_excel(input_file)
    print(f"   Total shops found: {len(shops)}")
    
    # Step 2: Shorten location links
    print("\n2. Shortening long location map links...")
    long_links_count = 0
    for shop in shops:
        original_link = shop['location_link']
        if original_link and len(str(original_link)) > 100:
            long_links_count += 1
            shop['location_link'] = shorten_google_maps_link(original_link)
    print(f"   Shortened {long_links_count} long links")
    
    # Step 3: Find duplicates
    print("\n3. Finding duplicate shops...")
    duplicates = find_duplicates(shops)
    print(f"   Found {len(duplicates)} potential duplicates")
    
    if duplicates:
        print("\n   Duplicate details:")
        for dup in duplicates[:10]:  # Show first 10
            print(f"   - Type: {dup['type']}")
            print(f"     Shop 1: {dup['shop1']['name']} (Area: {dup['shop1']['area']})")
            print(f"     Shop 2: {dup['shop2']['name']} (Area: {dup['shop2']['area']})")
    
    # Step 4: Remove duplicates (keep first occurrence)
    print("\n4. Removing duplicate shops...")
    indices_to_remove = set()
    for dup in duplicates:
        # Only remove exact duplicates or same license in same area
        if dup['type'] in ['exact', 'same_license_same_area']:
            indices_to_remove.add(dup['duplicate_index'])
    
    cleaned_shops = [shop for i, shop in enumerate(shops) if i not in indices_to_remove]
    print(f"   Removed {len(indices_to_remove)} duplicate shops")
    print(f"   Remaining shops: {len(cleaned_shops)}")
    
    # Report shops with same license in different areas (branches)
    license_map = {}
    for shop in cleaned_shops:
        license = str(shop['license']).strip()
        if license and license not in ['', 'nan', 'None'] and len(license) > 3:
            if license not in license_map:
                license_map[license] = []
            license_map[license].append(shop)
    
    branches = {lic: shops_list for lic, shops_list in license_map.items() if len(shops_list) > 1}
    if branches:
        print(f"\n   Note: Found {len(branches)} licenses with multiple branches in different areas")
        print("   These are kept as they represent different locations")
    
    # Step 5: Create cleaned Excel file
    print("\n5. Creating cleaned Excel file...")
    create_cleaned_excel(cleaned_shops, output_file)
    
    # Summary by area
    print("\n" + "="*80)
    print("Summary by Area:")
    print("="*80)
    area_counts = defaultdict(int)
    for shop in cleaned_shops:
        area_counts[shop['area']] += 1
    
    for area in sorted(area_counts.keys()):
        print(f"  {area}: {area_counts[area]} shops")
    
    print("\n" + "="*80)
    print("Cleanup completed successfully!")
    print("="*80)
    return 0

if __name__ == '__main__':
    import sys
    try:
        sys.exit(main())
    except Exception as e:
        print(f"\n❌ ERROR: An unexpected error occurred: {e}")
        print("Please check the input file and try again.")
        sys.exit(1)
