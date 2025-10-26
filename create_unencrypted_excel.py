#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Script to create an unencrypted Excel file from shops_details.json
This replaces the encrypted pet-shop-list-updated.xlsx file
"""

import json
import traceback
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

def create_excel_from_json(input_file='shops_details.json', output_file='pet-shop-list-updated.xlsx'):
    """Create an unencrypted Excel file from shops_details.json
    
    Args:
        input_file: Path to the JSON file containing shop data (default: 'shops_details.json')
        output_file: Path to the output Excel file (default: 'pet-shop-list-updated.xlsx')
    """
    
    # Load the JSON data
    print(f"Loading shops data from {input_file}...")
    with open(input_file, 'r', encoding='utf-8') as f:
        shops_data = json.load(f)
    
    print(f"Loaded {len(shops_data)} shops")
    
    # Create a new workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "قائمة المحلات"
    
    # Define headers
    headers = [
        "اسم المحل (عربي)",
        "Shop Name (English)",
        "رقم الرخصة",
        "العنوان",
        "رمز ADM",
        "معلومات الاتصال",
        "النشاط",
        "رابط الموقع"
    ]
    
    # Style for headers
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=12)
    
    # Write headers
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Write data rows
    row = 2
    for shop_name, shop_data in shops_data.items():
        ws.cell(row=row, column=1, value=shop_data.get('nameAr', shop_name))
        ws.cell(row=row, column=2, value=shop_data.get('nameEn', ''))
        ws.cell(row=row, column=3, value=shop_data.get('licenseNo', ''))
        ws.cell(row=row, column=4, value=shop_data.get('address', ''))
        ws.cell(row=row, column=5, value=shop_data.get('admCode', ''))
        ws.cell(row=row, column=6, value=shop_data.get('contact', ''))
        ws.cell(row=row, column=7, value=shop_data.get('activity', ''))
        ws.cell(row=row, column=8, value=shop_data.get('locationMap', ''))
        row += 1
    
    # Adjust column widths
    column_widths = {
        'A': 30,  # اسم المحل (عربي)
        'B': 30,  # Shop Name (English)
        'C': 15,  # رقم الرخصة
        'D': 25,  # العنوان
        'E': 12,  # رمز ADM
        'F': 25,  # معلومات الاتصال
        'G': 20,  # النشاط
        'H': 50   # رابط الموقع
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Freeze the header row
    ws.freeze_panes = "A2"
    
    # Save the file
    print(f"\nSaving to {output_file}...")
    wb.save(output_file)
    
    print(f"✅ Successfully created unencrypted Excel file!")
    print(f"   - Total shops: {len(shops_data)}")
    print(f"   - File: {output_file}")
    print(f"   - Size: {len(shops_data)} rows + 1 header row")
    print(f"\nThe file can now be opened in Excel without a password!")
    
    return output_file

if __name__ == "__main__":
    try:
        create_excel_from_json()
    except Exception as e:
        print(f"❌ Error: {e}")
        traceback.print_exc()
