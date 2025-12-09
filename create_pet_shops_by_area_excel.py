#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Create a comprehensive Excel workbook with pet shops organized by area.
Each area gets its own sheet with detailed shop information.
"""

import json
from collections import defaultdict
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def load_data():
    """Load data from plan-data.json and shops_details.json"""
    print("Loading data files...")
    
    # Load plan-data.json
    with open('plan-data.json', 'r', encoding='utf-8') as f:
        plan_data = json.load(f)
    
    # Load shops_details.json if exists
    shops_details = {}
    try:
        with open('shops_details.json', 'r', encoding='utf-8') as f:
            shops_details = json.load(f)
        print(f"✓ Loaded {len(shops_details)} shop details")
    except FileNotFoundError:
        print("⚠ shops_details.json not found, using basic info only")
    
    return plan_data, shops_details

def extract_shops_by_area(plan_data, shops_details):
    """Extract shops organized by area with full details"""
    print("\nExtracting shops by area...")
    
    # Initialize all areas from plan_data
    all_areas = set()
    if 'areas' in plan_data:
        for area in plan_data['areas']:
            if isinstance(area, dict):
                area_name = area.get('name', '')
                if area_name:
                    all_areas.add(area_name)
            else:
                all_areas.add(str(area))
    
    area_shops = defaultdict(list)
    shop_to_area = {}
    
    if 'inspectionData' in plan_data:
        inspection_data = plan_data['inspectionData']
    else:
        inspection_data = plan_data
    
    # Collect shops by area from inspection data
    for entry in inspection_data:
        area = entry.get('area', 'غير محدد')
        all_areas.add(area)  # Also add areas found in inspection data
        shops = entry.get('shops', [])
        
        for shop in shops:
            if shop not in shop_to_area:
                shop_to_area[shop] = area
                
                # Get detailed info if available
                shop_info = shops_details.get(shop, {})
                
                shop_data = {
                    'name_ar': shop_info.get('nameAr', shop),
                    'name_en': shop_info.get('nameEn', ''),
                    'license': shop_info.get('licenseNo', ''),
                    'address': shop_info.get('address', ''),
                    'adm_code': shop_info.get('admCode', ''),
                    'contact': shop_info.get('contact', ''),
                    'activity': shop_info.get('activity', ''),
                    'location_map': shop_info.get('locationMap', ''),
                    'area': area
                }
                
                area_shops[area].append(shop_data)
    
    # Also check shops_details for any shops with area information
    for shop_name, shop_info in shops_details.items():
        if 'area' in shop_info and shop_name not in shop_to_area:
            area = shop_info['area']
            all_areas.add(area)
            
            shop_data = {
                'name_ar': shop_info.get('nameAr', shop_name),
                'name_en': shop_info.get('nameEn', ''),
                'license': shop_info.get('licenseNo', ''),
                'address': shop_info.get('address', ''),
                'adm_code': shop_info.get('admCode', ''),
                'contact': shop_info.get('contact', ''),
                'activity': shop_info.get('activity', ''),
                'location_map': shop_info.get('locationMap', ''),
                'area': area
            }
            
            area_shops[area].append(shop_data)
            shop_to_area[shop_name] = area
    
    # Ensure all defined areas are in the dictionary (even if empty)
    for area in all_areas:
        if area not in area_shops:
            area_shops[area] = []
    
    # Sort shops within each area alphabetically
    for area in area_shops:
        area_shops[area].sort(key=lambda x: x['name_ar'])
    
    print(f"✓ Organized {len(shop_to_area)} shops across {len(area_shops)} areas")
    print(f"✓ Found {len(all_areas)} total areas (including areas with no shops)")
    
    return dict(area_shops)

def create_summary_sheet(wb, area_shops):
    """Create a summary sheet with statistics"""
    print("\nCreating summary sheet...")
    
    ws = wb.active
    ws.title = "📊 الملخص - Summary"
    
    # Title
    title_cell = ws.merge_cells('A1:F2')
    ws['A1'] = "قائمة محلات الحيوانات الأليفة حسب المناطق"
    ws['A1'].font = Font(bold=True, size=18, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells('A3:F3')
    ws['A3'] = "Pet Shops List by Areas"
    ws['A3'].font = Font(bold=True, size=14, color="666666")
    ws['A3'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Date
    ws.merge_cells('A4:F4')
    ws['A4'] = f"تاريخ الإنشاء: {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    ws['A4'].font = Font(size=10, italic=True)
    ws['A4'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers for area statistics
    row = 6
    headers = ["#", "المنطقة - Area", "عدد المحلات", "النسبة المئوية", "الحالة", "الملاحظات"]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=11)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Calculate statistics
    total_shops = sum(len(shops) for shops in area_shops.values())
    sorted_areas = sorted(area_shops.items(), key=lambda x: len(x[1]), reverse=True)
    
    # Write area statistics
    row = 7
    for idx, (area, shops) in enumerate(sorted_areas, start=1):
        shop_count = len(shops)
        percentage = (shop_count / total_shops * 100) if total_shops > 0 else 0
        
        # Determine status based on count
        if shop_count >= 30:
            status = "🔴 كبير جداً"
            status_color = "FFE6E6"
        elif shop_count >= 15:
            status = "🟡 كبير"
            status_color = "FFF4E6"
        elif shop_count >= 8:
            status = "🟢 متوسط"
            status_color = "E6F4EA"
        else:
            status = "🔵 صغير"
            status_color = "E6F2FF"
        
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=area)
        ws.cell(row=row, column=3, value=shop_count)
        ws.cell(row=row, column=4, value=f"{percentage:.1f}%")
        ws.cell(row=row, column=5, value=status)
        ws.cell(row=row, column=6, value=f"يحتوي على {shop_count} محل")
        
        # Apply formatting
        for col in range(1, 7):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            if col == 5:
                cell.fill = PatternFill(start_color=status_color, end_color=status_color, fill_type="solid")
        
        row += 1
    
    # Total row
    ws.cell(row=row, column=1, value="")
    ws.cell(row=row, column=2, value="المجموع الكلي - TOTAL")
    ws.cell(row=row, column=2).font = Font(bold=True, size=11)
    ws.cell(row=row, column=3, value=total_shops)
    ws.cell(row=row, column=3).font = Font(bold=True, size=11)
    ws.cell(row=row, column=4, value="100%")
    ws.cell(row=row, column=4).font = Font(bold=True, size=11)
    ws.cell(row=row, column=5, value="✅ مكتمل")
    ws.cell(row=row, column=5).font = Font(bold=True, size=11)
    ws.cell(row=row, column=6, value=f"{len(area_shops)} منطقة")
    ws.cell(row=row, column=6).font = Font(bold=True, size=11)
    
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col)
        cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='medium'),
            bottom=Side(style='medium')
        )
    
    # Add notes section
    row += 2
    ws.merge_cells(f'A{row}:F{row}')
    ws.cell(row=row, column=1, value="📌 ملاحظات هامة - Important Notes")
    ws.cell(row=row, column=1).font = Font(bold=True, size=12, color="FFFFFF")
    ws.cell(row=row, column=1).fill = PatternFill(start_color="70AD47", end_color="70AD47", fill_type="solid")
    ws.cell(row=row, column=1).alignment = Alignment(horizontal="center", vertical="center")
    
    notes = [
        "• كل منطقة لها صفحة مستقلة تحتوي على تفاصيل المحلات",
        "• Each area has its own sheet with shop details",
        "• البيانات مستخرجة من نظام خطة التفتيش الشهرية",
        "• Data extracted from Monthly Inspection Plan system"
    ]
    
    for note in notes:
        row += 1
        ws.merge_cells(f'A{row}:F{row}')
        ws.cell(row=row, column=1, value=note)
        ws.cell(row=row, column=1).alignment = Alignment(horizontal="right", vertical="center")
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 30
    
    print(f"✓ Summary sheet created with {len(area_shops)} areas and {total_shops} total shops")

def create_area_sheet(wb, area_name, shops):
    """Create a sheet for a specific area"""
    # Clean sheet name (Excel has limitations)
    sheet_name = area_name[:31]  # Max 31 characters
    ws = wb.create_sheet(title=sheet_name)
    
    # Title
    ws.merge_cells('A1:H2')
    ws['A1'] = f"محلات منطقة: {area_name}"
    ws['A1'].font = Font(bold=True, size=16, color="FFFFFF")
    ws['A1'].fill = PatternFill(start_color="2E5090", end_color="2E5090", fill_type="solid")
    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
    
    ws.merge_cells('A3:H3')
    ws['A3'] = f"Shops in Area: {area_name} | عدد المحلات: {len(shops)}"
    ws['A3'].font = Font(bold=True, size=11, color="666666")
    ws['A3'].alignment = Alignment(horizontal="center", vertical="center")
    
    # Headers
    row = 5
    headers = [
        "#",
        "اسم المحل\nShop Name",
        "رقم الرخصة\nLicense No",
        "العنوان\nAddress",
        "رمز ADM\nADM Code",
        "الاتصال\nContact",
        "النشاط\nActivity",
        "رابط الموقع\nLocation Map"
    ]
    
    for col, header in enumerate(headers, start=1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.font = Font(bold=True, color="FFFFFF", size=10)
        cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # Set row height for headers
    ws.row_dimensions[row].height = 35
    
    # Write shop data
    row = 6
    for idx, shop in enumerate(shops, start=1):
        ws.cell(row=row, column=1, value=idx)
        ws.cell(row=row, column=2, value=shop['name_ar'])
        ws.cell(row=row, column=3, value=shop['license'])
        ws.cell(row=row, column=4, value=shop['address'])
        ws.cell(row=row, column=5, value=shop['adm_code'])
        ws.cell(row=row, column=6, value=shop['contact'])
        ws.cell(row=row, column=7, value=shop['activity'])
        ws.cell(row=row, column=8, value=shop['location_map'])
        
        # Apply formatting
        for col in range(1, 9):
            cell = ws.cell(row=row, column=col)
            cell.alignment = Alignment(horizontal="right" if col in [2, 4, 7] else "center", 
                                     vertical="center", 
                                     wrap_text=True)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Alternate row colors
            if idx % 2 == 0:
                cell.fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        row += 1
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 30
    ws.column_dimensions['C'].width = 15
    ws.column_dimensions['D'].width = 25
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 20
    ws.column_dimensions['H'].width = 40
    
    # Freeze panes
    ws.freeze_panes = "A6"
    
    print(f"  ✓ Created sheet for '{area_name}' with {len(shops)} shops")

def create_excel_workbook(area_shops, output_file='pet-shops-by-area.xlsx'):
    """Create the complete Excel workbook"""
    print("\n" + "="*80)
    print("CREATING PET SHOPS EXCEL WORKBOOK BY AREA")
    print("="*80)
    
    wb = Workbook()
    
    # Create summary sheet
    create_summary_sheet(wb, area_shops)
    
    # Create sheets for each area (sorted by name)
    print("\nCreating area sheets...")
    for area_name in sorted(area_shops.keys()):
        shops = area_shops[area_name]
        create_area_sheet(wb, area_name, shops)
    
    # Save workbook
    print(f"\n💾 Saving workbook to '{output_file}'...")
    wb.save(output_file)
    
    # Summary
    total_shops = sum(len(shops) for shops in area_shops.values())
    print("\n" + "="*80)
    print("✅ EXCEL WORKBOOK CREATED SUCCESSFULLY!")
    print("="*80)
    print(f"📊 Summary:")
    print(f"   • Total areas: {len(area_shops)}")
    print(f"   • Total shops: {total_shops}")
    print(f"   • Total sheets: {len(area_shops) + 1} (1 summary + {len(area_shops)} area sheets)")
    print(f"   • Output file: {output_file}")
    print("="*80)
    
    return output_file

def main():
    """Main function"""
    try:
        # Load data
        plan_data, shops_details = load_data()
        
        # Extract shops by area
        area_shops = extract_shops_by_area(plan_data, shops_details)
        
        # Create Excel workbook
        output_file = create_excel_workbook(area_shops)
        
        print(f"\n✨ The Excel file '{output_file}' is ready to use!")
        print("   You can open it in Microsoft Excel, Google Sheets, or any spreadsheet software.")
        
    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback
        traceback.print_exc()
        return 1
    
    return 0

if __name__ == '__main__':
    exit(main())
