#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Smart merge script for pet shop licenses
Merges data from 'رخص المحلات الجديدة.xlsx' into 'pet-shops.xlsx' and 'shops_details.json'
Based on specific keywords related to pet shops, animals, birds, and fish.
"""

import openpyxl
import json
import sys
from datetime import datetime

# Keywords to filter relevant shops (from problem statement)
KEYWORDS = [
    'الطيور', 'الأسماك', 'تجارة الأسماك', 'اكسسوارات', 
    'بيع اغذية', 'إيواء', 'بيع اكسسوارات', 'تجارة الحيوانات', 
    'تجارة الطيور', 'تربية الصقور', 'بيع الأسماك', 
    'صالون حلاقة', 'صالون', 'علاج بيطري', 'عيادة بيطرية', 
    'استيراد', 'أغذية', 'الحيوانات'
]

def contains_keyword(text, keywords):
    """Check if text contains any of the keywords"""
    if not text:
        return False
    text_str = str(text)
    for keyword in keywords:
        if keyword in text_str:
            return True
    return False

def extract_area_name(address):
    """
    Extract simplified area name from detailed address.
    
    Examples:
        "شمال الوثبة, شمال الوثبة 59, ق 10 - مستودع 31" -> "شمال الوثبة"
        "المصفح, م 43, 0 : ~, مبنى" -> "المصفح"
        "أبو ظبي - شارع الميناء - بناية" -> "أبو ظبي"
        "الوثبة, مسلخ الوثبة - زريبة 39" -> "الوثبة"
        "مصفح جنوب, ايكاد 3, 0 : ~" -> "مصفح جنوب"
        " أبو ظبي - شارع  الميناء -  بناية دائرة بلدية أبوظبي, " -> "أبو ظبي"
    
    Strategy:
        1. Strip whitespace and trailing comma
        2. If address contains comma, take the first part before comma
        3. If no comma but contains dash, take first part before dash
        4. Strip and return
    """
    if not address:
        return ''
    
    # Strip whitespace and trailing commas
    address_str = str(address).strip().rstrip(',').strip()
    
    # If address contains comma, extract first part
    if ',' in address_str:
        area = address_str.split(',')[0].strip()
        return area
    
    # If address contains dash, extract first part
    if '-' in address_str:
        area = address_str.split('-')[0].strip()
        return area
    
    # Return the address as is (already simple)
    return address_str

def extract_new_licenses():
    """Extract relevant rows from new licenses Excel file"""
    print("📖 Reading 'رخص المحلات الجديدة.xlsx'...")
    wb = openpyxl.load_workbook('رخص المحلات الجديدة.xlsx', data_only=True)
    ws = wb.active
    
    new_shops = []
    for i in range(2, ws.max_row + 1):  # Start from row 2 (skip header)
        row_data = {
            'A': ws[f'A{i}'].value,  # LICENSE_NO
            'B': ws[f'B{i}'].value,  # TRADE_NAME_AR
            'C': ws[f'C{i}'].value,  # TRADE_NAME_EN
            'M': ws[f'M{i}'].value,  # EMAIL
            'Q': ws[f'Q{i}'].value,  # ADDRESS
            'S': ws[f'S{i}'].value,  # ACTIVITY_NAME_AR
            'T': ws[f'T{i}'].value,  # ACTIVITY_NAME_EN
        }
        
        # Check if any keyword appears in any of the specified columns
        matched = False
        for col_value in row_data.values():
            if contains_keyword(col_value, KEYWORDS):
                matched = True
                break
        
        if matched:
            new_shops.append(row_data)
    
    print(f"✅ Found {len(new_shops)} shops matching keywords")
    return new_shops

def load_existing_shops():
    """Load existing shops from pet-shops.xlsx"""
    print("📖 Reading existing 'pet-shops.xlsx'...")
    wb = openpyxl.load_workbook('pet-shops.xlsx', data_only=True)
    ws = wb.active
    
    existing_licenses = set()
    for i in range(2, ws.max_row + 1):  # Skip header
        license_no = ws[f'A{i}'].value
        if license_no:
            existing_licenses.add(str(license_no).strip())
    
    print(f"✅ Found {len(existing_licenses)} existing shops")
    return wb, ws, existing_licenses

def merge_to_excel(new_shops, existing_licenses):
    """Merge new shops into pet-shops.xlsx"""
    wb, ws, existing = load_existing_shops()
    
    added_count = 0
    skipped_count = 0
    last_row = ws.max_row
    
    for shop in new_shops:
        license_no = str(shop['A']).strip() if shop['A'] else ''
        
        # Skip if already exists
        if license_no in existing:
            skipped_count += 1
            continue
        
        # Add new row with simplified area name
        last_row += 1
        ws[f'A{last_row}'] = shop['A']  # LICENSE_NO
        ws[f'B{last_row}'] = shop['B']  # TRADE_NAME_AR
        ws[f'C{last_row}'] = shop['C']  # TRADE_NAME_EN
        ws[f'M{last_row}'] = shop['M']  # EMAIL
        ws[f'Q{last_row}'] = extract_area_name(shop['Q'])  # ADDRESS (simplified)
        ws[f'S{last_row}'] = shop['S']  # ACTIVITY_NAME_AR
        ws[f'T{last_row}'] = shop['T']  # ACTIVITY_NAME_EN
        
        added_count += 1
    
    # Save the updated Excel file
    wb.save('pet-shops.xlsx')
    print(f"✅ Added {added_count} new shops to pet-shops.xlsx")
    print(f"⏭️  Skipped {skipped_count} duplicate shops")
    
    return added_count

def load_shops_details():
    """Load shops_details.json"""
    print("📖 Reading 'shops_details.json'...")
    try:
        with open('shops_details.json', 'r', encoding='utf-8') as f:
            shops_details = json.load(f)
        print(f"✅ Found {len(shops_details)} shops in JSON")
        return shops_details
    except Exception as e:
        print(f"⚠️  Error reading shops_details.json: {e}")
        return {}

def generate_adm_code(existing_codes):
    """Generate next ADM code"""
    # Extract numeric parts from existing codes
    max_num = 0
    for code in existing_codes:
        if code and code.startswith('ADM'):
            try:
                num = int(code[3:])
                max_num = max(max_num, num)
            except:
                pass
    
    return f"ADM{max_num + 1:04d}"

def merge_to_json(new_shops, shops_details):
    """Merge new shops into shops_details.json"""
    existing_licenses = {shop.get('licenseNo') for shop in shops_details.values() if shop.get('licenseNo')}
    existing_codes = {shop.get('admCode') for shop in shops_details.values() if shop.get('admCode')}
    
    added_count = 0
    skipped_count = 0
    
    for shop in new_shops:
        license_no = str(shop['A']).strip() if shop['A'] else ''
        name_ar = str(shop['B']).strip() if shop['B'] else ''
        
        # Skip if already exists (by license number)
        if license_no in existing_licenses:
            skipped_count += 1
            continue
        
        # Skip if no name
        if not name_ar:
            continue
        
        # Create shop entry with simplified area name
        shop_entry = {
            'nameAr': name_ar,
            'nameEn': str(shop['C']).strip() if shop['C'] else name_ar,
            'licenseNo': license_no,
            'locationMap': '',  # Will be empty for new shops
            'admCode': generate_adm_code(existing_codes),
            'address': extract_area_name(shop['Q']),  # Simplified area name
            'contact': str(shop['M']).strip() if shop['M'] else '',
            'activity': str(shop['S']).strip() if shop['S'] else ''
        }
        
        # Add to shops_details with Arabic name as key
        shops_details[name_ar] = shop_entry
        existing_codes.add(shop_entry['admCode'])
        existing_licenses.add(license_no)
        added_count += 1
    
    # Save updated JSON
    with open('shops_details.json', 'w', encoding='utf-8') as f:
        json.dump(shops_details, f, ensure_ascii=False, indent=2)
    
    print(f"✅ Added {added_count} new shops to shops_details.json")
    print(f"⏭️  Skipped {skipped_count} duplicate shops in JSON")
    
    return added_count

def main():
    """Main merge process"""
    print("="*80)
    print("🔄 Smart Pet Shop Licenses Merge")
    print("="*80)
    print(f"Started at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    print()
    
    try:
        # Step 1: Extract new shops from licenses file
        new_shops = extract_new_licenses()
        if not new_shops:
            print("⚠️  No shops found matching keywords")
            return 1
        
        print()
        
        # Step 2: Merge into pet-shops.xlsx
        excel_added = merge_to_excel(new_shops, set())
        
        print()
        
        # Step 3: Merge into shops_details.json
        shops_details = load_shops_details()
        json_added = merge_to_json(new_shops, shops_details)
        
        print()
        print("="*80)
        print("✨ Merge Completed Successfully!")
        print("="*80)
        print(f"Total shops added to Excel: {excel_added}")
        print(f"Total shops added to JSON: {json_added}")
        print(f"Completed at: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        return 0
        
    except Exception as e:
        print()
        print("="*80)
        print("❌ Error during merge:")
        print("="*80)
        print(str(e))
        import traceback
        traceback.print_exc()
        return 1

if __name__ == '__main__':
    sys.exit(main())
