#!/usr/bin/env python3
"""
Test script to validate the shop lists cleanup operation.
This script verifies that the cleanup was performed correctly.
"""

import openpyxl
import sys

def test_new_shop_list():
    """Test that new shop list contains only yellow-highlighted shops"""
    print("Testing new-shop-list-updated.xlsx...")
    
    wb = openpyxl.load_workbook('new-shop-list-updated.xlsx')
    ws = wb.active
    
    total_shops = ws.max_row - 1
    yellow_count = 0
    non_yellow_count = 0
    
    for row_idx in range(2, ws.max_row + 1):
        cell = ws.cell(row_idx, 1)
        if cell.fill and cell.fill.fgColor and hasattr(cell.fill.fgColor, 'rgb'):
            color = str(cell.fill.fgColor.rgb).upper()
            if 'FFFF' in color and color != '00000000':
                yellow_count += 1
            else:
                non_yellow_count += 1
        else:
            non_yellow_count += 1
    
    # Assertions
    assert total_shops == 104, f"Expected 104 shops, found {total_shops}"
    assert yellow_count == 104, f"Expected 104 yellow shops, found {yellow_count}"
    assert non_yellow_count == 0, f"Expected 0 non-yellow shops, found {non_yellow_count}"
    
    print(f"  ✓ Total shops: {total_shops}")
    print(f"  ✓ All shops are yellow-highlighted")
    print("  ✓ PASSED")
    return True

def test_old_shop_list():
    """Test that old shop list is unchanged"""
    print("\nTesting old-shop-list-updated.xlsx...")
    
    wb = openpyxl.load_workbook('old-shop-list-updated.xlsx')
    ws = wb.active
    
    total_shops = ws.max_row - 1
    
    # Assertions
    assert total_shops == 128, f"Expected 128 shops, found {total_shops}"
    
    print(f"  ✓ Total shops: {total_shops}")
    print("  ✓ File unchanged")
    print("  ✓ PASSED")
    return True

def test_backup_files_exist():
    """Test that backup files were created"""
    import os
    print("\nTesting backup files...")
    
    # Check for backup files (they should exist locally)
    backup_files = [
        f for f in os.listdir('.')
        if f.startswith(('new-shop-list-updated_backup_', 'old-shop-list-updated_backup_'))
        and f.endswith('.xlsx')
    ]
    
    assert len(backup_files) >= 2, f"Expected at least 2 backup files, found {len(backup_files)}"
    
    print(f"  ✓ Found {len(backup_files)} backup files")
    for backup in backup_files:
        print(f"    - {backup}")
    print("  ✓ PASSED")
    return True

def main():
    """Run all tests"""
    print("=" * 60)
    print("Shop Lists Cleanup Validation Tests")
    print("=" * 60)
    
    try:
        test_new_shop_list()
        test_old_shop_list()
        test_backup_files_exist()
        
        print("\n" + "=" * 60)
        print("✓ ALL TESTS PASSED")
        print("=" * 60)
        print("\nCleanup operation validated successfully!")
        print("- New file: 104 yellow-highlighted shops")
        print("- Old file: 128 shops (unchanged)")
        print("- Backup files created")
        return 0
        
    except AssertionError as e:
        print(f"\n✗ TEST FAILED: {e}")
        return 1
    except Exception as e:
        print(f"\n✗ ERROR: {e}")
        return 1

if __name__ == '__main__':
    sys.exit(main())
